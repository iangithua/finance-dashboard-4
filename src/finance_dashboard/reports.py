from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage

from .analytics import (
    build_dashboard_metrics,
    monthly_cash_flow,
    category_summary,
    money_leak_report,
    top_transactions,
    unclassified_transactions,
)
from .charts import generate_charts

def _format_sheet(writer, sheet_name: str):
    ws = writer.book[sheet_name]
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for column in ws.columns:
        max_len = 0
        col_letter = column[0].column_letter
        for cell in column:
            max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

def _embed_image(ws, image_path: str, cell: str, width: int = 640):
    if not image_path or not Path(image_path).exists():
        return
    img = XLImage(image_path)
    ratio = img.height / img.width if img.width else 0.6
    img.width = width
    img.height = int(width * ratio)
    ws.add_image(img, cell)

def export_reports(df: pd.DataFrame, settings: dict) -> dict:
    excel_dir = Path(settings["paths"]["output_excel_dir"])
    csv_dir = Path(settings["paths"]["output_csv_dir"])
    reports_dir = Path(settings["paths"]["output_reports_dir"])
    charts_dir = Path(settings["paths"].get("output_charts_dir", "output/charts"))

    excel_dir.mkdir(parents=True, exist_ok=True)
    csv_dir.mkdir(parents=True, exist_ok=True)
    reports_dir.mkdir(parents=True, exist_ok=True)
    charts_dir.mkdir(parents=True, exist_ok=True)

    small_threshold = float(settings["reports"]["small_purchase_threshold"])
    top_limit = int(settings["reports"]["top_transactions_limit"])

    dashboard = build_dashboard_metrics(df, small_threshold)
    monthly = monthly_cash_flow(df)
    categories = category_summary(df)
    leaks = money_leak_report(df, small_threshold)
    top = top_transactions(df, top_limit)
    unclassified = unclassified_transactions(df)

    chart_paths = generate_charts(monthly, categories, leaks, charts_dir)

    excel_path = excel_dir / "finance_dashboard.xlsx"
    csv_path = csv_dir / "finance_dashboard_summary.csv"

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        dashboard.to_excel(writer, sheet_name="Dashboard", index=False)
        monthly.to_excel(writer, sheet_name="Monthly Cash Flow", index=False)
        categories.to_excel(writer, sheet_name="Category Summary", index=False)
        leaks.to_excel(writer, sheet_name="Money Leak Report", index=False)
        top.to_excel(writer, sheet_name="Top Transactions", index=False)
        unclassified.to_excel(writer, sheet_name="Unclassified", index=False)
        df.to_excel(writer, sheet_name="All Transactions", index=False)

        chart_ws = writer.book.create_sheet("Charts")
        chart_ws["A1"] = "Finance Charts"
        chart_ws["A1"].font = Font(bold=True, size=16)
        chart_ws["A3"] = "Monthly cash flow"
        chart_ws["A30"] = "Top spending categories"
        chart_ws["J3"] = "Spending by group"
        chart_ws["J30"] = "Money leaks"
        for cell in ["A3", "A30", "J3", "J30"]:
            chart_ws[cell].font = Font(bold=True)

        _embed_image(chart_ws, chart_paths.get("monthly_cash_flow"), "A4", width=560)
        _embed_image(chart_ws, chart_paths.get("top_spending_categories"), "A31", width=560)
        _embed_image(chart_ws, chart_paths.get("spending_by_group"), "J4", width=520)
        _embed_image(chart_ws, chart_paths.get("money_leaks"), "J31", width=560)

        dash_ws = writer.book["Dashboard"]
        _embed_image(dash_ws, chart_paths.get("monthly_cash_flow"), "D2", width=620)

        for sheet_name in writer.book.sheetnames:
            if sheet_name != "Charts":
                _format_sheet(writer, sheet_name)

    dashboard.to_csv(csv_path, index=False)
    monthly.to_csv(reports_dir / "monthly_cash_flow.csv", index=False)
    categories.to_csv(reports_dir / "category_summary.csv", index=False)
    leaks.to_csv(reports_dir / "money_leak_report.csv", index=False)
    unclassified.to_csv(reports_dir / "unclassified_transactions.csv", index=False)
    df.to_csv(reports_dir / "all_classified_transactions.csv", index=False)

    return {
        "excel": str(excel_path),
        "csv": str(csv_path),
        "reports_dir": str(reports_dir),
        "charts_dir": str(charts_dir),
        "charts": chart_paths,
    }
